#!/usr/bin/env python3
"""
genera_psc_completo.py
Aggiorna tutti i 9 PSC DOCX con:
  1. Foto reali dal foto.pdf di ogni UI (sostituisce i 11 slot JPEG nel template)
  2. Gantt chart generato per ogni cantiere (image2.png)
  3. Tabella valutazione rischi per ogni cantiere (image15.png)
  4. File Excel adattati (AnalisiRischi, Cronoprogramma, OneriSicurezza) embedded come OLE

Uso: python genera_psc_completo.py

Dipendenze:
    pip install python-docx pymupdf openpyxl matplotlib numpy requests
"""

import os, sys, re, shutil, io, zipfile, requests, textwrap
from pathlib import Path
from copy import deepcopy
from datetime import datetime, timedelta

# ── Importazioni con verifica ─────────────────────────────────────────────────
try:
    import fitz                          # PyMuPDF
except ImportError:
    sys.exit("pip install pymupdf")

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    import numpy as np
except ImportError:
    sys.exit("pip install matplotlib numpy")

try:
    from lxml import etree
except ImportError:
    sys.exit("pip install lxml")

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURAZIONE PERCORSI
# ══════════════════════════════════════════════════════════════════════════════
ANTHROPIC_KEY = os.environ.get("ANTHROPIC_KEY", "")  # imposta la variabile d'ambiente ANTHROPIC_KEY
MODEL_CLAUDE  = "claude-haiku-4-5-20251001"

# Su Windows
BASE       = Path(r"C:\Users\franz\OneDrive\Desktop\PSC ITEA per Frenz")
DATI_DIR   = BASE / "03 Progetto e dati ITEA" / "trento_doss trento"
XLS_DIR    = BASE / "02 PSC e XLS di partenza"
OUTPUT_DIR = Path(r"C:\Users\franz\safetyai\PSC_OUTPUT")

# Rileva se siamo su Linux (container) e adatta i percorsi
import platform
if platform.system() == "Linux":
    BASE       = Path("/sessions/sleepy-sharp-edison/mnt/03 Progetto e dati ITEA").parent.parent / \
                 "03 Progetto e dati ITEA"
    # Ricostruisci percorsi per Linux
    MNT        = Path("/sessions/sleepy-sharp-edison/mnt")
    BASE       = MNT / "03 Progetto e dati ITEA"
    DATI_DIR   = BASE / "trento_doss trento"
    XLS_DIR    = MNT / "02 PSC e XLS di partenza"
    OUTPUT_DIR = MNT / "safetyai" / "PSC_OUTPUT"

TEMP_DIR = Path("/tmp/psc_temp")
TEMP_DIR.mkdir(exist_ok=True)

# ══════════════════════════════════════════════════════════════════════════════
# DATI DEI 9 CANTIERI
# ══════════════════════════════════════════════════════════════════════════════
def piano_str(p):
    return {"1":"1°","2":"2°","3":"3°","4":"4°","5":"5°","6":"6°"}.get(str(p), f"{p}°")

def fmt_euro(v):
    intero, dec = f"{v:.2f}".split(".")
    gruppi = []
    while len(intero) > 3:
        gruppi.insert(0, intero[-3:])
        intero = intero[:-3]
    gruppi.insert(0, intero)
    return ".".join(gruppi) + "," + dec

def gg_to_mesi(gg):
    m = round(gg / 30)
    return f"{m} mesi"

MESI_IT = ["", "gennaio","febbraio","marzo","aprile","maggio","giugno",
           "luglio","agosto","settembre","ottobre","novembre","dicembre"]

MESI_IT_UPPER = [m.upper() for m in MESI_IT]

def parse_data_inizio(s):
    """'8 giugno 2026' → datetime(2026,6,8)"""
    parts = s.split()
    g = int(parts[0])
    m = MESI_IT.index(parts[1].lower())
    a = int(parts[2])
    return datetime(a, m, g)

CANTIERI = [
    {
        "codice_ui"      : "43103914",
        "cod_ed"         : "958",
        "civico"         : "45/1",
        "scala"          : "A",
        "interno"        : "23",
        "piano"          : "5",
        "p_ed"           : "5620",
        "sup"            : "99,81",
        "anno_costr"     : "1983",
        "amianto"        : False,
        "importo"        : 17534.71,
        "oneri"          : 3109.31,
        "oneri_lettere"  : "tremilacentonove/31",
        "durata_gg"      : 120,
        "data_inizio"    : "8 giugno 2026",
        "tipo_intervento": "sistemazione parziale",
    },
    {
        "codice_ui"      : "43104709",
        "cod_ed"         : "1346",
        "civico"         : "37",
        "scala"          : "A",
        "interno"        : "8",
        "piano"          : "2",
        "p_ed"           : "2496/3",
        "sup"            : "37,29",
        "anno_costr"     : "1988",
        "amianto"        : False,
        "importo"        : 32056.73,
        "oneri"          : 3500.00,
        "oneri_lettere"  : "tremilasettecento/00",
        "durata_gg"      : 120,
        "data_inizio"    : "22 giugno 2026",
        "tipo_intervento": "sistemazione completa",
    },
    {
        "codice_ui"      : "43104715",
        "cod_ed"         : "1346",
        "civico"         : "37",
        "scala"          : "A",
        "interno"        : "14",
        "piano"          : "3",
        "p_ed"           : "2496/3",
        "sup"            : "37,29",
        "anno_costr"     : "1988",
        "amianto"        : False,
        "importo"        : 31717.14,
        "oneri"          : 3500.00,
        "oneri_lettere"  : "tremilasettecento/00",
        "durata_gg"      : 120,
        "data_inizio"    : "6 luglio 2026",
        "tipo_intervento": "sistemazione completa",
    },
    {
        "codice_ui"      : "43105028",
        "cod_ed"         : "1421",
        "civico"         : "35",
        "scala"          : "A",
        "interno"        : "4",
        "piano"          : "1",
        "p_ed"           : "2496/6",
        "sup"            : "28,59",
        "anno_costr"     : "1989",
        "amianto"        : False,
        "importo"        : 9067.72,
        "oneri"          : 2000.00,
        "oneri_lettere"  : "duemila/00",
        "durata_gg"      : 60,
        "data_inizio"    : "20 luglio 2026",
        "tipo_intervento": "sistemazione parziale",
    },
    {
        "codice_ui"      : "43105031",
        "cod_ed"         : "1421",
        "civico"         : "35",
        "scala"          : "A",
        "interno"        : "7",
        "piano"          : "2",
        "p_ed"           : "2496/3",
        "sup"            : "78,55",
        "anno_costr"     : "1989",
        "amianto"        : False,
        "importo"        : 42198.01,
        "oneri"          : 4000.00,
        "oneri_lettere"  : "quattromila/00",
        "durata_gg"      : 140,
        "data_inizio"    : "3 agosto 2026",
        "tipo_intervento": "sistemazione completa",
    },
    {
        "codice_ui"      : "43105049",
        "cod_ed"         : "1421",
        "civico"         : "35",
        "scala"          : "A",
        "interno"        : "25",
        "piano"          : "5",
        "p_ed"           : "2496/6",
        "sup"            : "78,54",
        "anno_costr"     : "1989",
        "amianto"        : False,
        "importo"        : 46801.38,
        "oneri"          : 4200.00,
        "oneri_lettere"  : "quattromiladuecento/00",
        "durata_gg"      : 140,
        "data_inizio"    : "17 agosto 2026",
        "tipo_intervento": "sistemazione completa",
    },
    {
        "codice_ui"      : "43105448",
        "cod_ed"         : "1487",
        "civico"         : "31",
        "scala"          : "A",
        "interno"        : "5",
        "piano"          : "1",
        "p_ed"           : "6047",
        "sup"            : "78,38",
        "anno_costr"     : "1990",
        "amianto"        : True,
        "importo"        : 48277.21,
        "oneri"          : 5500.00,
        "oneri_lettere"  : "cinquemilacinquecento/00",
        "durata_gg"      : 160,
        "data_inizio"    : "31 agosto 2026",
        "tipo_intervento": "ristrutturazione parziale",
    },
    {
        "codice_ui"      : "43105459",
        "cod_ed"         : "1487",
        "civico"         : "31",
        "scala"          : "A",
        "interno"        : "16",
        "piano"          : "4",
        "p_ed"           : "6047",
        "sup"            : "77,31",
        "anno_costr"     : "1990",
        "amianto"        : False,
        "importo"        : 46068.67,
        "oneri"          : 4200.00,
        "oneri_lettere"  : "quattromiladuecento/00",
        "durata_gg"      : 140,
        "data_inizio"    : "14 settembre 2026",
        "tipo_intervento": "ristrutturazione parziale",
    },
    {
        "codice_ui"      : "43105460",
        "cod_ed"         : "1487",
        "civico"         : "31",
        "scala"          : "A",
        "interno"        : "17",
        "piano"          : "4",
        "p_ed"           : "6047",
        "sup"            : "46,23",
        "anno_costr"     : "1990",
        "amianto"        : False,
        "importo"        : 33434.18,
        "oneri"          : 3600.00,
        "oneri_lettere"  : "tremilaseicento/00",
        "durata_gg"      : 100,
        "data_inizio"    : "28 settembre 2026",
        "tipo_intervento": "sistemazione completa",
    },
]

# ══════════════════════════════════════════════════════════════════════════════
# TEMPLATE TASK LISTS PER TIPO INTERVENTO
# ══════════════════════════════════════════════════════════════════════════════
# (task_name, start_week, duration_weeks, color, categoria)
TASKS_COMPLETO = [
    ("Allestimento cantiere",              0,  1, "#4472C4", "Coordinamento"),
    ("Demolizioni e rimozioni interne",    1,  2, "#4472C4", "Coordinamento"),
    ("Nuova rete gas interna",             2,  2, "#FF0000", "Ditta gas"),
    ("Rifacimento impianto elettrico",     2,  5, "#FFC000", "Ditta elettrica"),
    ("Sostituzione caldaia a condensaz.",  3,  2, "#ED7D31", "Ditta impianti"),
    ("Sostituzione radiatori + tubaz.",    3,  3, "#ED7D31", "Ditta impianti"),
    ("Box doccia + rubinetteria",          6,  2, "#ED7D31", "Ditta impianti"),
    ("Sostituzione sanitari",              7,  2, "#ED7D31", "Ditta impianti"),
    ("Sostituzione porte interne",         6,  2, "#70AD47", "Ditta serramenti"),
    ("Revisione serramenti esterni",       8,  2, "#70AD47", "Ditta serramenti"),
    ("Pavimenti caldi (ceramica)",         6,  4, "#9DC3E6", "Ditta pavimenti"),
    ("Pavimenti freddi (ceramica)",        7,  3, "#9DC3E6", "Ditta pavimenti"),
    ("Sostituzione citofono",             10,  1, "#FFC000", "Ditta elettrica"),
    ("Tinteggiatura alloggio",            11,  4, "#A9D18E", "Ditta tinteggiatura"),
    ("Verniciatura parapetto",            13,  2, "#A9D18E", "Ditta tinteggiatura"),
    ("Pulizia e smontaggio cantiere",     15,  1, "#BFBFBF", "Coordinamento"),
]

TASKS_PARZIALE = [
    ("Allestimento cantiere",              0,  1, "#4472C4", "Coordinamento"),
    ("Demolizioni parziali",               1,  1, "#4472C4", "Coordinamento"),
    ("Rifacimento impianto elettrico",     1,  4, "#FFC000", "Ditta elettrica"),
    ("Sostituzione caldaia",               2,  2, "#ED7D31", "Ditta impianti"),
    ("Sostituzione sanitari",              4,  2, "#ED7D31", "Ditta impianti"),
    ("Revisione serramenti",               4,  2, "#70AD47", "Ditta serramenti"),
    ("Tinteggiatura alloggio",             6,  3, "#A9D18E", "Ditta tinteggiatura"),
    ("Pulizia e smontaggio cantiere",      7,  1, "#BFBFBF", "Coordinamento"),
]

TASKS_RISTRUTTURAZIONE = [
    ("Allestimento cantiere",              0,  1, "#4472C4", "Coordinamento"),
    ("Bonifica amianto (se presente)",     0,  2, "#FF0000", "Ditta bonifica"),
    ("Demolizioni e rimozioni",            2,  2, "#4472C4", "Coordinamento"),
    ("Nuova rete gas interna",             3,  2, "#FF0000", "Ditta gas"),
    ("Rifacimento impianto elettrico",     3,  5, "#FFC000", "Ditta elettrica"),
    ("Sostituzione caldaia + radiatori",   4,  3, "#ED7D31", "Ditta impianti"),
    ("Impianto idraulico + sanitari",      6,  3, "#ED7D31", "Ditta impianti"),
    ("Sostituzione porte e serramenti",    7,  3, "#70AD47", "Ditta serramenti"),
    ("Pavimenti (ceramica e parquet)",     8,  4, "#9DC3E6", "Ditta pavimenti"),
    ("Controsoffitti e cartongessi",       9,  3, "#BDD7EE", "Ditta edile"),
    ("Tinteggiatura e rivestimenti",      12,  4, "#A9D18E", "Ditta tinteggiatura"),
    ("Verniciatura parapetto",            15,  2, "#A9D18E", "Ditta tinteggiatura"),
    ("Pulizia e smontaggio cantiere",     17,  1, "#BFBFBF", "Coordinamento"),
]

def get_tasks(tipo):
    if "completa" in tipo:
        return TASKS_COMPLETO
    elif "ristrutturazione" in tipo:
        return TASKS_RISTRUTTURAZIONE
    else:
        return TASKS_PARZIALE

# ══════════════════════════════════════════════════════════════════════════════
# GENERAZIONE GANTT PNG
# ══════════════════════════════════════════════════════════════════════════════
def genera_gantt_png(c: dict) -> Path:
    ui    = c["codice_ui"]
    civ   = c["civico"]
    sc    = c["scala"]
    inn   = c["interno"]
    pn    = piano_str(c["piano"])
    gg    = c["durata_gg"]
    tipo  = c["tipo_intervento"]
    di    = c["data_inizio"]

    tasks     = get_tasks(tipo)
    # Calcola le settimane (30 min → 8 sett. min; 160 gg → ~23 sett.)
    n_weeks   = max(8, round(gg / 7) + 1)
    # Adatta durate proporzionalmente se necessario
    max_end_week = max(t[1] + t[2] for t in tasks)
    if max_end_week > n_weeks - 1:
        scale = (n_weeks - 1) / max_end_week
        tasks = [(nm, round(sw*scale), max(1, round(dw*scale)), col, cat)
                 for nm, sw, dw, col, cat in tasks]

    # Calcola etichette mesi partendo da data_inizio
    dt0    = parse_data_inizio(di)
    months_seen = []
    month_positions = []
    for w in range(n_weeks + 1):
        dt_w = dt0 + timedelta(weeks=w)
        label = f"{MESI_IT_UPPER[dt_w.month]} {dt_w.year}"
        if not months_seen or months_seen[-1] != label:
            months_seen.append(label)
            month_positions.append(w)

    fig, ax = plt.subplots(figsize=(16, max(7, len(tasks)*0.55 + 2.5)))
    fig.patch.set_facecolor('white')

    for i in range(n_weeks + 1):
        ax.axvline(x=i, color='#CCCCCC', linewidth=0.5, zorder=0)
    for i in range(len(tasks) + 1):
        ax.axhline(y=i, color='#CCCCCC', linewidth=0.5, zorder=0)

    for idx, (name, start, dur, color, cat) in enumerate(tasks):
        y = len(tasks) - 1 - idx
        ax.barh(y, dur, left=start, height=0.65, color=color,
                edgecolor='white', linewidth=0.8, zorder=2)
        if dur >= 2:
            ax.text(start + dur/2, y, name, va='center', ha='center',
                    fontsize=7, color='black', fontweight='bold', zorder=3)
        else:
            ax.text(start + dur + 0.1, y, name, va='center', ha='left',
                    fontsize=7, color='black', zorder=3)

    ax.set_yticks(range(len(tasks)))
    ax.set_yticklabels([t[0] for t in reversed(tasks)], fontsize=8)
    ax.set_xlim(0, n_weeks)
    ax.set_xticks(range(n_weeks + 1))
    week_lbls = [''] + [f"S.{i+1}" for i in range(n_weeks)]
    ax.set_xticklabels(week_lbls, fontsize=7, rotation=45, ha='right')

    ax2 = ax.twiny()
    ax2.set_xlim(0, n_weeks)
    ax2.set_xticks(month_positions)
    ax2.set_xticklabels(months_seen, fontsize=8, color='#2F5496')

    ax.set_title(
        f"CRONOPROGRAMMA LAVORI – UI {ui}\n"
        f"Via Doss Trento {civ}, Sc. {sc}, Int. {inn}, P. {pn} – {tipo.title()} – {gg} giorni",
        fontsize=11, fontweight='bold', pad=15, color='#1F3864')

    # Legenda categorie
    cats_seen = {}
    for _, _, _, col, cat in tasks:
        cats_seen[cat] = col
    patches = [mpatches.Patch(color=c, label=l) for l, c in cats_seen.items()]
    ax.legend(handles=patches, loc='lower right', fontsize=7.5,
              framealpha=0.9, edgecolor='#CCCCCC')

    plt.tight_layout()
    out = TEMP_DIR / f"gantt_{ui}.png"
    plt.savefig(str(out), dpi=200, bbox_inches='tight', facecolor='white')
    plt.close()
    return out


# ══════════════════════════════════════════════════════════════════════════════
# GENERAZIONE TABELLA RISCHI PNG
# ══════════════════════════════════════════════════════════════════════════════
def get_lavorazioni_rischi(tipo: str, amianto: bool) -> list:
    base = [
        (1,  "Allestimento impianto di cantiere",      2, 2, 4,  "#FFFF00"),
        (2,  "Demolizioni e rimozioni",                2, 2, 4,  "#FFFF00"),
    ]
    if amianto:
        base.insert(1, (999, "Bonifica amianto (ditta specializzata)", 2, 4, 8, "#FFC000"))

    completo = [
        (3,  "Nuova rete gas interna",                 2, 4, 8,  "#FFC000"),
        (4,  "Rifacimento impianto elettrico",         2, 3, 6,  "#FFC000"),
        (5,  "Sostituzione citofono",                  1, 2, 2,  "#92D050"),
        (6,  "Sostituzione caldaia a condensazione",   2, 3, 6,  "#FFC000"),
        (7,  "Sostituzione radiatori e tubazioni",     2, 2, 4,  "#FFFF00"),
        (8,  "Sostituzione porte interne",             2, 2, 4,  "#FFFF00"),
        (9,  "Revisione serramenti esterni",           1, 2, 2,  "#92D050"),
        (10, "Rifacimento pavimenti caldi (ceramica)", 2, 2, 4,  "#FFFF00"),
        (11, "Rifacimento pavimenti freddi",           2, 2, 4,  "#FFFF00"),
        (12, "Sostituzione sanitari e rubinetteria",   2, 2, 4,  "#FFFF00"),
        (13, "Impianto idraulico – box doccia",        2, 2, 4,  "#FFFF00"),
        (14, "Tinteggiatura alloggio",                 2, 2, 4,  "#FFFF00"),
        (15, "Verniciatura parapetto e poggiolo",      2, 3, 6,  "#FFC000"),
        (16, "Smaltimento materiali di risulta",       2, 2, 4,  "#FFFF00"),
        (17, "Pulizia finale alloggio",                1, 2, 2,  "#92D050"),
        (18, "Smontaggio impianto di cantiere",        1, 1, 1,  "#E2EFDA"),
    ]
    parziale = [
        (3,  "Rifacimento impianto elettrico",         2, 3, 6,  "#FFC000"),
        (4,  "Sostituzione caldaia",                   2, 3, 6,  "#FFC000"),
        (5,  "Sostituzione sanitari",                  2, 2, 4,  "#FFFF00"),
        (6,  "Revisione serramenti",                   1, 2, 2,  "#92D050"),
        (7,  "Tinteggiatura alloggio",                 2, 2, 4,  "#FFFF00"),
        (8,  "Smaltimento materiali di risulta",       2, 2, 4,  "#FFFF00"),
        (9,  "Pulizia finale alloggio",                1, 1, 1,  "#E2EFDA"),
        (10, "Smontaggio impianto di cantiere",        1, 1, 1,  "#E2EFDA"),
    ]
    ristrutturazione = [
        (3,  "Nuova rete gas interna",                 2, 4, 8,  "#FFC000"),
        (4,  "Rifacimento impianto elettrico",         2, 3, 6,  "#FFC000"),
        (5,  "Sostituzione caldaia + radiatori",       2, 3, 6,  "#FFC000"),
        (6,  "Impianto idraulico + sanitari",          2, 2, 4,  "#FFFF00"),
        (7,  "Sostituzione porte e serramenti",        2, 2, 4,  "#FFFF00"),
        (8,  "Pavimenti (ceramica e parquet)",         2, 2, 4,  "#FFFF00"),
        (9,  "Controsoffitti e cartongessi",           2, 3, 6,  "#FFC000"),
        (10, "Tinteggiatura e rivestimenti",           2, 2, 4,  "#FFFF00"),
        (11, "Verniciatura parapetto",                 2, 3, 6,  "#FFC000"),
        (12, "Smaltimento materiali di risulta",       2, 2, 4,  "#FFFF00"),
        (13, "Pulizia finale e smontaggio cantiere",   1, 1, 1,  "#E2EFDA"),
    ]
    if "completa" in tipo:
        rows = base + completo
    elif "ristrutturazione" in tipo:
        rows = base + ristrutturazione
    else:
        rows = base + parziale
    # Rinumera
    for i, row in enumerate(rows):
        rows[i] = (i+1,) + row[1:]
    return rows


def genera_rischi_png(c: dict) -> Path:
    ui   = c["codice_ui"]
    civ  = c["civico"]
    sc   = c["scala"]
    inn  = c["interno"]
    tipo = c["tipo_intervento"]

    lavorazioni = get_lavorazioni_rischi(tipo, c.get("amianto", False))

    risk_labels = {1: "Accettabile", 2: "Basso", 4: "Moderato",
                   6: "Medio", 8: "Medio-Alto", 12: "Alto", 16: "Molto Alto"}

    import matplotlib.patches as patches

    n        = len(lavorazioni)
    col_widths = [0.5, 5.5, 1.0, 1.0, 1.0, 1.5]
    headers  = ["N.", "Lavorazione / Fase", "P", "D", "R", "Livello di Rischio"]
    total_w  = sum(col_widths)
    row_h    = 0.42
    fig_h    = (n + 2) * row_h + 1.4

    fig, ax = plt.subplots(figsize=(14, fig_h))
    fig.patch.set_facecolor('white')
    ax.set_xlim(0, total_w)
    ax.set_ylim(0, fig_h)
    ax.axis('off')

    ax.text(total_w/2, fig_h - 0.35,
            f"VALUTAZIONE DEI RISCHI – UI {ui}",
            ha='center', va='center', fontsize=11, fontweight='bold', color='#1F3864')
    ax.text(total_w/2, fig_h - 0.70,
            f"Via Doss Trento {civ}, Scala {sc}, Int. {inn} – Trento – {tipo.title()}",
            ha='center', va='center', fontsize=8, color='#444444')

    def draw_cell(x, y, w, h, text, bg='white', fontsize=8, bold=False, center=True):
        rect = patches.FancyBboxPatch((x, y), w, h,
                                       boxstyle="square,pad=0",
                                       linewidth=0.5, edgecolor='#888888',
                                       facecolor=bg, zorder=2)
        ax.add_patch(rect)
        ha  = 'center' if center else 'left'
        px  = x + w/2 if center else x + 0.06
        txt = ax.text(px, y + h/2, text, ha=ha, va='center',
                      fontsize=fontsize, fontweight='bold' if bold else 'normal',
                      zorder=3, clip_on=True)
        return txt

    header_y = fig_h - 1.05
    x = 0
    for hdr, cw in zip(headers, col_widths):
        draw_cell(x, header_y - row_h, cw, row_h, hdr,
                  bg='#2F5496', fontsize=8, bold=True)
        ax.texts[-1].set_color('white')
        x += cw

    for i, (num, desc, p, d, r, fill) in enumerate(lavorazioni):
        row_y  = header_y - row_h - (i+1) * row_h
        row_bg = '#F2F2F2' if i % 2 == 0 else 'white'
        x = 0
        cells    = [str(num), desc, str(p), str(d), str(r),
                    risk_labels.get(r, f"R={r}")]
        bgs      = [row_bg]*5 + [fill]
        centers  = [True, False, True, True, True, True]
        for col, (txt, cw) in enumerate(zip(cells, col_widths)):
            draw_cell(x, row_y, cw, row_h, txt,
                      bg=bgs[col], fontsize=7.5, bold=(col==0), center=centers[col])
            x += cw

    legend_y = header_y - row_h - (n+1)*row_h - 0.05
    ax.text(0, legend_y,
            "P = Probabilità (1=improbabile, 2=poco prob., 3=probabile, 4=molto prob.)   "
            "D = Danno (1=lieve, 2=medio, 3=grave, 4=gravissimo)   R = P × D",
            fontsize=6.5, color='#555555', va='top')

    plt.tight_layout(pad=0.3)
    out = TEMP_DIR / f"rischi_{ui}.png"
    plt.savefig(str(out), dpi=200, bbox_inches='tight', facecolor='white')
    plt.close()
    return out


# ══════════════════════════════════════════════════════════════════════════════
# ESTRAZIONE FOTO DAL PDF
# ══════════════════════════════════════════════════════════════════════════════

# I 11 slot JPEG nel template (nome file → numero progressivo nel PDF)
JPEG_SLOTS = [
    "image24.jpeg", "image25.jpeg", "image26.jpeg",
    "image28.jpeg", "image29.jpeg", "image31.jpeg",
    "image34.jpeg", "image36.jpeg", "image37.jpeg",
    "image38.jpeg", "image44.jpg",
]

def estrai_foto(codice_ui: str) -> list[Path]:
    """
    Estrae le foto dal foto.pdf (20 pagine, 2 img/pag).
    Restituisce una lista di percorsi PNG (max 11, una per ogni slot JPEG).
    """
    ui_dir = DATI_DIR / f"UI_{codice_ui}"
    candidati = [
        ui_dir / f"UI_{codice_ui}_foto.pdf",
        ui_dir / f"UI_{codice_ui}_foto.PDF",
    ]
    pdf_path = next((p for p in candidati if p.exists()), None)
    if not pdf_path:
        print(f"  ⚠  foto.pdf non trovato per UI_{codice_ui}")
        return []

    doc    = fitz.open(str(pdf_path))
    n_pag  = doc.page_count
    foto   = []
    # Estrai la prima immagine di ogni pagina (quelle dispari, che tipicamente
    # contengono la foto vera e propria piuttosto che il modulo)
    for pg_idx in range(n_pag):
        if len(foto) >= len(JPEG_SLOTS):
            break
        page = doc[pg_idx]
        imgs = page.get_images(full=True)
        if not imgs:
            continue
        # Prendi la prima immagine embedded
        xref = imgs[0][0]
        try:
            pix = fitz.Pixmap(doc, xref)
            if pix.n < 3:       # scala di grigi → converti in RGB
                pix = fitz.Pixmap(fitz.csRGB, pix)
            out_path = TEMP_DIR / f"foto_{codice_ui}_{pg_idx+1:02d}.jpg"
            # Se è già JPEG, scrivi direttamente; altrimenti converti
            raw = doc.extract_image(xref)
            ext = raw.get("ext", "png")
            if ext in ("jpeg", "jpg"):
                out_path.write_bytes(raw["image"])
            else:
                pix.save(str(out_path.with_suffix(".jpg")))
                out_path = out_path.with_suffix(".jpg")
            foto.append(out_path)
        except Exception as e:
            print(f"  ⚠  Errore estrazione img pg {pg_idx+1}: {e}")

    doc.close()

    if len(foto) < len(JPEG_SLOTS):
        # Completa ciclicamente se ci sono meno foto dei slot
        while len(foto) < len(JPEG_SLOTS) and foto:
            foto.append(foto[len(foto) % len(foto)])

    print(f"  📷  Estratte {len(foto)} foto da {pdf_path.name}")
    return foto[:len(JPEG_SLOTS)]


# ══════════════════════════════════════════════════════════════════════════════
# ADATTAMENTO EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def calcola_mesi_gantt(data_inizio: str, durata_gg: int) -> list[str]:
    """Calcola i 4 nomi mesi (con anno) da inserire nelle celle C5,G5,K5,O5"""
    dt0    = parse_data_inizio(data_inizio)
    mesi   = []
    seen   = set()
    for w in range(20):
        dt_w = dt0 + timedelta(weeks=w)
        label = f"{MESI_IT_UPPER[dt_w.month]} {dt_w.year}"
        if label not in seen:
            seen.add(label)
            mesi.append(label)
        if len(mesi) == 4:
            break
    while len(mesi) < 4:
        last = mesi[-1]
        m, y = last.split()
        mi = MESI_IT_UPPER.index(m)
        mi = mi % 12 + 1
        y = int(y) + (1 if mi == 1 else 0)
        mesi.append(f"{MESI_IT_UPPER[mi]} {y}")
    return mesi


def adatta_analisi_rischi(c: dict) -> Path:
    ui   = c["codice_ui"]
    civ  = c["civico"]
    sc   = c["scala"]
    inn  = c["interno"]
    tipo = c["tipo_intervento"]

    src  = XLS_DIR / "AnalisiRischi_UI_43103914.xlsx"
    dst  = TEMP_DIR / f"AnalisiRischi_UI_{ui}.xlsx"
    shutil.copy(src, dst)

    wb = openpyxl.load_workbook(str(dst))
    ws = wb['RISCHI']
    # Aggiorna descrizione cantiere
    ws['B2'] = (f"Cantiere UI {ui} - {tipo.title()} - "
                f"via Doss Trento {civ}, sc.{sc}, int. {inn} - TRENTO")
    wb.save(str(dst))
    wb.close()
    return dst


def adatta_cronoprogramma(c: dict) -> Path:
    ui   = c["codice_ui"]
    civ  = c["civico"]
    sc   = c["scala"]
    inn  = c["interno"]
    gg   = c["durata_gg"]
    mesi_n = round(gg / 30)
    di   = c["data_inizio"]

    src  = XLS_DIR / "Cronoprogramma_UI_43103914.xlsx"
    dst  = TEMP_DIR / f"Cronoprogramma_UI_{ui}.xlsx"
    shutil.copy(src, dst)

    wb  = openpyxl.load_workbook(str(dst))
    ws  = wb['GANTT']

    ws['B2'] = (f"Lavori di manutenzione ordinaria e straordinaria - "
                f"UI {ui} - via Doss Trento {civ}, scala {sc}, interno {inn}, TRENTO")
    ws['B3'] = (f"Inizio presunto: {di} - "
                f"Durata complessiva: {mesi_n} mesi ({round(gg/7)} settimane)")

    mesi4 = calcola_mesi_gantt(di, gg)
    ws['C5'] = mesi4[0]
    ws['G5'] = mesi4[1]
    ws['K5'] = mesi4[2]
    ws['O5'] = mesi4[3]

    wb.save(str(dst))
    wb.close()
    return dst


def adatta_oneri(c: dict) -> Path:
    ui   = c["codice_ui"]
    civ  = c["civico"]
    sc   = c["scala"]
    inn  = c["interno"]
    pn   = piano_str(c["piano"])
    gg   = c["durata_gg"]
    mesi_n = round(gg / 30)
    di   = c["data_inizio"]
    imp  = fmt_euro(c["importo"])
    on   = fmt_euro(c["oneri"])

    dt0  = parse_data_inizio(di)
    dt_f = dt0 + timedelta(days=gg)
    fine = f"{dt_f.day} {MESI_IT[dt_f.month]} {dt_f.year}"

    src  = XLS_DIR / "OneriSicurezza_UI_43103914.xlsx"
    dst  = TEMP_DIR / f"OneriSicurezza_UI_{ui}.xlsx"
    shutil.copy(src, dst)

    wb = openpyxl.load_workbook(str(dst))
    ws = wb['Stima Oneri Sicurezza']
    ws['B3'] = (f"Cantiere: UI_{ui} – Via Doss Trento {civ}, "
                f"scala {sc}, interno {inn}, piano {pn} – Trento")
    ws['B4'] = (f"Committente: ITEA S.p.A. – "
                f"Durata lavori: {mesi_n} mesi ({di} – {fine}) – CME: € {imp}")

    wb.save(str(dst))
    wb.close()
    return dst


# ══════════════════════════════════════════════════════════════════════════════
# GENERAZIONE ANTEPRIMA PNG DEGLI EXCEL (via LibreOffice headless)
# ══════════════════════════════════════════════════════════════════════════════
def genera_preview_excel(xlsx_path: Path) -> Path:
    """Converte il primo foglio dell'Excel in PNG via LibreOffice."""
    import subprocess
    result = subprocess.run(
        ["libreoffice", "--headless", "--convert-to", "png",
         str(xlsx_path), "--outdir", str(TEMP_DIR)],
        capture_output=True, text=True, timeout=60
    )
    png_path = TEMP_DIR / (xlsx_path.stem + ".png")
    if png_path.exists():
        print(f"  📊  Preview: {png_path.name} ({png_path.stat().st_size//1024} KB)")
        return png_path
    # Fallback: genera preview con matplotlib
    print(f"  ⚠  LibreOffice preview fallito per {xlsx_path.name}, uso fallback")
    return genera_preview_fallback(xlsx_path)


def genera_preview_fallback(xlsx_path: Path) -> Path:
    """Crea una preview minimale dell'Excel come immagine."""
    wb  = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws  = wb.active
    rows_data = []
    for row in ws.iter_rows(min_row=1, max_row=min(30, ws.max_row), values_only=True):
        rows_data.append([str(c)[:40] if c else '' for c in row[:8]])
    wb.close()

    fig, ax = plt.subplots(figsize=(14, 7))
    fig.patch.set_facecolor('white')
    ax.axis('off')
    ax.set_title(xlsx_path.stem, fontsize=12, fontweight='bold', pad=10)

    table = ax.table(cellText=rows_data, loc='center', cellLoc='left')
    table.auto_set_font_size(False)
    table.set_fontsize(8)
    table.scale(1, 1.3)

    out = TEMP_DIR / (xlsx_path.stem + ".png")
    plt.savefig(str(out), dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    return out


# ══════════════════════════════════════════════════════════════════════════════
# SOSTITUZIONE IMMAGINI NEL DOCX (manipolazione ZIP)
# ══════════════════════════════════════════════════════════════════════════════
def sostituisci_immagine_in_docx(docx_path: Path, img_name: str,
                                  new_img_path: Path) -> bool:
    """
    Sostituisce word/media/{img_name} nel DOCX con il contenuto di new_img_path.
    Restituisce True se la sostituzione è avvenuta.
    """
    target = f"word/media/{img_name}"
    tmp    = docx_path.with_suffix(".tmp.docx")

    with zipfile.ZipFile(docx_path, 'r') as zin, \
         zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename == target:
                zout.writestr(item.filename, new_img_path.read_bytes())
            else:
                zout.writestr(item, zin.read(item.filename))

    tmp.replace(docx_path)
    return True


# ══════════════════════════════════════════════════════════════════════════════
# EMBEDDING OLE DEGLI EXCEL NEL DOCX
# ══════════════════════════════════════════════════════════════════════════════
_OLE_SHAPE_COUNTER = 1000

def embed_excel_in_docx(docx_path: Path,
                         xlsx_path: Path,
                         preview_png: Path,
                         anchor_text: str) -> bool:
    """
    Embeds an Excel file as OLE object in a DOCX, after the paragraph
    containing anchor_text.
    """
    global _OLE_SHAPE_COUNTER
    _OLE_SHAPE_COUNTER += 1
    shape_id  = _OLE_SHAPE_COUNTER
    obj_id    = f"_{shape_id * 1000}"

    # Leggi il DOCX in memoria
    with zipfile.ZipFile(docx_path, 'r') as z:
        files = {name: z.read(name) for name in z.namelist()}

    # ── 1. Aggiungi l'Excel a word/embeddings/ ────────────────────────────
    emb_name  = f"word/embeddings/{xlsx_path.name}"
    files[emb_name] = xlsx_path.read_bytes()

    # ── 2. Aggiungi la preview PNG a word/media/ ──────────────────────────
    prev_name = f"word/media/preview_{xlsx_path.stem}.png"
    files[prev_name] = preview_png.read_bytes()

    # ── 3. Aggiungi relazioni in document.xml.rels ────────────────────────
    rels_key  = "word/_rels/document.xml.rels"
    rels_xml  = files[rels_key].decode('utf-8')

    # Trova l'Id massimo esistente
    existing_ids = re.findall(r'Id="rId(\d+)"', rels_xml)
    max_id = max((int(x) for x in existing_ids), default=60)
    rid_embed   = f"rId{max_id + 1}"
    rid_preview = f"rId{max_id + 2}"

    new_rels = (
        f'<Relationship Id="{rid_embed}" '
        f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/oleObject" '
        f'Target="embeddings/{xlsx_path.name}"/>\n'
        f'<Relationship Id="{rid_preview}" '
        f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
        f'Target="media/preview_{xlsx_path.stem}.png"/>\n'
    )
    rels_xml = rels_xml.replace('</Relationships>', new_rels + '</Relationships>')
    files[rels_key] = rels_xml.encode('utf-8')

    # ── 4. Aggiungi content type per l'Excel ─────────────────────────────
    ct_key = "[Content_Types].xml"
    ct_xml = files[ct_key].decode('utf-8')
    ct_entry = (
        f'<Override PartName="/word/embeddings/{xlsx_path.name}" '
        f'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"/>'
    )
    if xlsx_path.name not in ct_xml:
        ct_xml = ct_xml.replace('</Types>', ct_entry + '\n</Types>')
        files[ct_key] = ct_xml.encode('utf-8')

    # ── 5. Inserisci il blocco OLE nel document.xml ───────────────────────
    doc_xml = files["word/document.xml"].decode('utf-8')

    # Calcola dimensioni preview in EMU (English Metric Units)
    from PIL import Image as PILImage
    try:
        with PILImage.open(str(preview_png)) as img:
            pw, ph = img.size
    except Exception:
        pw, ph = 1920, 1080

    # Target width: ~16 cm = 5760960 EMU; altezza proporzionale
    target_w_emu = 5760960
    target_h_emu = int(target_w_emu * ph / pw)
    # In pt (1 pt = 12700 EMU); dxa = pt * 20
    w_pt = target_w_emu // 12700
    h_pt = target_h_emu // 12700
    w_dxa = w_pt * 20
    h_dxa = h_pt * 20

    ole_xml = f"""<w:p xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"
             xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
             xmlns:v="urn:schemas-microsoft-com:vml"
             xmlns:o="urn:schemas-microsoft-com:office:office">
  <w:pPr><w:jc w:val="left"/></w:pPr>
  <w:r>
    <w:object w:dxaOrig="{w_dxa}" w:dyaOrig="{h_dxa}">
      <v:shape id="{shape_id}" o:spid="_x0000_i{shape_id}" type="#_x0000_t75"
               style="width:{w_pt}pt;height:{h_pt}pt" o:ole="" fillcolor="window">
        <v:imagedata r:id="{rid_preview}" o:title="{xlsx_path.stem}"/>
      </v:shape>
      <o:OLEObject Type="Embed" ProgID="Excel.Sheet.12" ShapeID="{shape_id}"
                   DrawAspect="Content" ObjectID="{obj_id}" r:id="{rid_embed}"/>
    </w:object>
  </w:r>
</w:p>"""

    # Trova il paragrafo con l'anchor_text e inserisci dopo di esso
    # Cerchiamo il testo nel documento XML senza parsing completo
    # L'anchor è il nome del file Excel (es: "AnalisiRischi_UI_43104709")
    anchor_key = xlsx_path.stem  # es. "AnalisiRischi_UI_43104709"
    # Cerca </w:p> dopo l'occorrenza dell'anchor_text
    idx = doc_xml.find(anchor_key)
    if idx == -1:
        # Prova con il nome generico (senza UI code)
        generic = xlsx_path.stem.split('_UI_')[0]
        idx = doc_xml.find(generic)

    if idx != -1:
        # Trova la fine del paragrafo che contiene l'anchor
        end_para = doc_xml.find('</w:p>', idx)
        if end_para != -1:
            insert_pos = end_para + len('</w:p>')
            doc_xml = doc_xml[:insert_pos] + "\n" + ole_xml + doc_xml[insert_pos:]
            print(f"  📎  OLE inserito dopo '{anchor_key[:40]}' [pos {idx}]")
        else:
            print(f"  ⚠  Fine paragrafo non trovata per '{anchor_key}'")
            return False
    else:
        # Appendi alla fine del body
        body_end = doc_xml.rfind('</w:body>')
        if body_end != -1:
            doc_xml = doc_xml[:body_end] + ole_xml + '\n' + doc_xml[body_end:]
            print(f"  📎  OLE appeso in fondo (anchor '{anchor_key}' non trovato)")

    files["word/document.xml"] = doc_xml.encode('utf-8')

    # ── 6. Riscrivi il DOCX ───────────────────────────────────────────────
    tmp = docx_path.with_suffix(".tmp.docx")
    with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)
    tmp.replace(docx_path)
    return True


# ══════════════════════════════════════════════════════════════════════════════
# ELABORAZIONE DI UN SINGOLO PSC
# ══════════════════════════════════════════════════════════════════════════════

# Dimensione immagine2 originale nel template (Gantt placeholder)
TEMPLATE_IMG2_SIZE  = 358565
TEMPLATE_IMG15_SIZE = 503278

def elabora_psc(c: dict) -> Path:
    ui  = c["codice_ui"]
    print(f"\n{'═'*64}")
    print(f"  UI_{ui}  —  Via Doss Trento {c['civico']} sc.{c['scala']} int.{c['interno']}")
    print(f"{'═'*64}")

    # ── Sorgente DOCX ─────────────────────────────────────────────────────
    bak = OUTPUT_DIR / f"PSC_UI_{ui}.bak.docx"
    src = OUTPUT_DIR / f"PSC_UI_{ui}.docx"
    if bak.exists():
        sorgente = bak
        print(f"  → Uso sorgente: {bak.name}")
    elif src.exists():
        sorgente = src
        print(f"  → Uso sorgente: {src.name}")
    else:
        print(f"  ✗  Nessun DOCX sorgente trovato per UI_{ui}")
        return None

    # Lavora su copia temporanea (evita SameFileError se src == dst)
    out_path  = OUTPUT_DIR / f"PSC_UI_{ui}.docx"
    work_path = TEMP_DIR   / f"PSC_UI_{ui}_work.docx"
    shutil.copy(str(sorgente), str(work_path))

    # ── Verifica se Gantt e Rischi vanno rigenerati ────────────────────────
    with zipfile.ZipFile(work_path, 'r') as z:
        try: sz_img2  = z.getinfo("word/media/image2.png").file_size
        except: sz_img2  = 0
        try: sz_img15 = z.getinfo("word/media/image15.png").file_size
        except: sz_img15 = 0

    rigenera_gantt  = (sz_img2  == TEMPLATE_IMG2_SIZE  or sz_img2  == 0)
    rigenera_rischi = (sz_img15 == TEMPLATE_IMG15_SIZE or sz_img15 == 0)

    # ── 1. Gantt chart ────────────────────────────────────────────────────
    if rigenera_gantt:
        print(f"  📈  Generazione Gantt…")
        gantt_png = genera_gantt_png(c)
        sostituisci_immagine_in_docx(work_path, "image2.png", gantt_png)
        print(f"  ✓  image2.png sostituita ({gantt_png.stat().st_size//1024} KB)")
    else:
        print(f"  ✓  Gantt già presente (image2.png = {sz_img2//1024} KB)")

    # ── 2. Tabella rischi ─────────────────────────────────────────────────
    if rigenera_rischi:
        print(f"  📋  Generazione tabella rischi…")
        rischi_png = genera_rischi_png(c)
        sostituisci_immagine_in_docx(work_path, "image15.png", rischi_png)
        print(f"  ✓  image15.png sostituita ({rischi_png.stat().st_size//1024} KB)")
    else:
        print(f"  ✓  Tabella rischi già presente (image15.png = {sz_img15//1024} KB)")

    # ── 3. Foto reali ─────────────────────────────────────────────────────
    print(f"  📷  Estrazione foto…")
    foto_paths = estrai_foto(ui)
    if foto_paths:
        for slot_name, foto_path in zip(JPEG_SLOTS, foto_paths):
            if foto_path.exists():
                sostituisci_immagine_in_docx(work_path, slot_name, foto_path)
        print(f"  ✓  {len(foto_paths)} foto sostituite ({', '.join(JPEG_SLOTS[:3])}…)")
    else:
        print(f"  ⚠  Nessuna foto estratta, slot JPEG lasciati invariati")

    # ── 4. Adatta Excel files ─────────────────────────────────────────────
    print(f"  📊  Adattamento file Excel…")
    xl_analisi    = adatta_analisi_rischi(c)
    xl_cronoprog  = adatta_cronoprogramma(c)
    xl_oneri      = adatta_oneri(c)
    print(f"  ✓  Excel adattati: {xl_analisi.name}, {xl_cronoprog.name}, {xl_oneri.name}")

    # ── 5. Preview PNG degli Excel ────────────────────────────────────────
    print(f"  🖼   Generazione anteprime Excel…")
    prev_analisi   = genera_preview_excel(xl_analisi)
    prev_cronoprog = genera_preview_excel(xl_cronoprog)
    prev_oneri     = genera_preview_excel(xl_oneri)

    # ── 6. Embedding OLE nel DOCX ─────────────────────────────────────────
    print(f"  📎  Embedding OLE…")
    embed_excel_in_docx(work_path, xl_analisi,   prev_analisi,   f"AnalisiRischi_UI_{ui}")
    embed_excel_in_docx(work_path, xl_cronoprog, prev_cronoprog, f"Cronoprogramma_UI_{ui}")
    embed_excel_in_docx(work_path, xl_oneri,     prev_oneri,     f"OneriSicurezza_UI_{ui}")

    # ── 7. Copia il DOCX finale in OUTPUT_DIR ────────────────────────────
    # Strategia: prova out_path; se bloccato, usa il file sorgente originale
    def _try_write(src_path: Path, dst_path: Path) -> bool:
        try:
            # Scrivi sul file destinazione già esistente (r+b per verificare)
            with open(str(dst_path), 'r+b') as _:
                pass
            # Il file è accessibile in scrittura: sovrascrivilo
            with open(str(src_path), 'rb') as fi, open(str(dst_path), 'wb') as fo:
                fo.write(fi.read())
            return True
        except (FileNotFoundError, PermissionError, OSError):
            return False

    salvato_in = None
    if out_path.exists() and _try_write(work_path, out_path):
        salvato_in = out_path
        print(f"  ✓  DOCX finale → {out_path.name}")
    elif sorgente.exists() and sorgente != out_path and _try_write(work_path, sorgente):
        salvato_in = sorgente
        print(f"  ✓  DOCX finale → {sorgente.name}  (out_path bloccato)")
    else:
        # Fallback: salva in TEMP_DIR (l'utente potrà copiarlo manualmente)
        fallback = TEMP_DIR / f"PSC_UI_{ui}_FINALE.docx"
        shutil.copy(str(work_path), str(fallback))
        salvato_in = fallback
        print(f"  ⚠  DOCX finale → {fallback}  (copiare manualmente in PSC_OUTPUT)")

    # ── 8. Copia Excel anche come file separati accanto al PSC ────────────
    for xl in [xl_analisi, xl_cronoprog, xl_oneri]:
        xl_dst = OUTPUT_DIR / xl.name
        if xl_dst.exists():
            try:
                with open(str(xl), 'rb') as fi, open(str(xl_dst), 'wb') as fo:
                    fo.write(fi.read())
            except Exception:
                shutil.copy(str(xl), str(TEMP_DIR / xl.name))
        else:
            # File nuovo: prova a crearlo
            try:
                with open(str(xl), 'rb') as fi, open(str(xl_dst), 'wb') as fo:
                    fo.write(fi.read())
            except Exception:
                shutil.copy(str(xl), str(TEMP_DIR / xl.name))
    print(f"  ✓  Excel salvati in PSC_OUTPUT (o in temp se non scrivibile)")

    sz = salvato_in.stat().st_size
    print(f"\n  ✅  UI_{ui} completato — {sz//1024} KB  [{salvato_in.name}]")
    return salvato_in


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════

if __name__ == "__main__":
    try:
        from PIL import Image
    except ImportError:
        import subprocess as _sp
        _sp.run([__import__('sys').executable, "-m", "pip", "install",
                 "Pillow", "--break-system-packages", "-q"])

    print()
    print("╔══════════════════════════════════════════════════════════════╗")
    print("║   Generazione PSC completi — ITEA Via Doss Trento, Trento   ║")
    print("║   Foto + Gantt + Rischi + Excel OLE per 9 cantieri          ║")
    print("╚══════════════════════════════════════════════════════════════╝")
    import sys as _sys
    print(f"  Output: {OUTPUT_DIR}", flush=True)
    print(f"  Temp  : {TEMP_DIR}", flush=True)
    print()

    completati, errori = [], []

    for cantiere in CANTIERI:
        try:
            out = elabora_psc(cantiere)
            if out:
                completati.append(out.name)
        except Exception as exc:
            import traceback
            print(f"\n  ✗  ERRORE per UI_{cantiere['codice_ui']}: {exc}", flush=True)
            traceback.print_exc()
            errori.append(cantiere["codice_ui"])

    print()
    print("══════════════════════════════════════════════════════════════")
    print(f"  Completati: {len(completati)} / {len(CANTIERI)}")
    for nome in completati:
        print(f"    ✅  {nome}")
    if errori:
        print(f"  Errori: UI_{', UI_'.join(errori)}")
    print(f"  Output → {OUTPUT_DIR}")
    print(f"  Temp   → {TEMP_DIR}")
    print("══════════════════════════════════════════════════════════════")
